"""Tabular exports of every artifact — CSV per dataset, or one multi-sheet XLSX.

instructions.txt: "Full export functionality" across the architecture, skills
taxonomy, task taxonomy and 3rd-party matches.

Each dataset is defined once as (columns, rows) and rendered to either format,
so a CSV and its XLSX sheet can never drift apart. The audit columns the
clustering methodology depends on — backbone vs final cluster, stability score,
whether an LLM moved it — are exported too: an architecture a client cannot
interrogate is not defensible, and those columns are the interrogation.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.project_state import ProjectState


@dataclass
class Dataset:
    name: str          # sheet name / file stem
    columns: list[str]
    rows: list[list]

    def to_csv(self) -> bytes:
        buf = io.StringIO(newline="")
        w = csv.writer(buf, lineterminator="\r\n")
        w.writerow(self.columns)
        w.writerows(self.rows)
        # utf-8-sig so Excel on Windows opens accented job titles correctly
        # rather than showing mojibake.
        return buf.getvalue().encode("utf-8-sig")


HEADER_FILL = PatternFill("solid", fgColor="1E2233")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def to_xlsx(datasets: list[Dataset]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for ds in datasets:
        # Excel sheet names cap at 31 chars and reject []:*?/\
        safe = ds.name[:31]
        ws = wb.create_sheet(safe)
        ws.append(ds.columns)
        for row in ds.rows:
            ws.append(row)

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for i, col in enumerate(ds.columns, start=1):
            longest = max(
                [len(str(col))] + [len(str(r[i - 1])) for r in ds.rows[:200] if i <= len(r)]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, longest + 2))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Dataset builders ─────────────────────────────────────────────────────────


def _profile_headcount(state: ProjectState) -> dict[str, int]:
    if not state.clustering:
        return {}
    hc = {r.id: r.headcount for r in state.raw_records}
    members = {g.group_id: g.member_ids for g in state.dedupe_groups}
    cluster_to_key = {d.profile_cluster_id: d.profile_key for d in state.job_profiles}
    out: dict[str, int] = {}
    for a in state.clustering.assignments:
        key = cluster_to_key.get(a.final_profile_id)
        if not key:
            continue
        total = sum(h for h in (hc.get(m) for m in members.get(a.item_id, [a.item_id])) if h)
        if total:
            out[key] = out.get(key, 0) + total
    return out


def architecture_dataset(state: ProjectState) -> Dataset:
    """One row per job profile — the headline deliverable."""
    c = state.clustering
    cols = [
        "Job family", "Job category", "Job profile", "Profile key", "Headcount",
        "Input jobs", "JE score", "JE level",
        "Taxonomy family", "Taxonomy specialization", "Taxonomy code",
        "Career level", "Match confidence", "Match needs review", "Match set by user",
        "Skills required", "Tasks", "Stale",
    ]
    if c is None:
        return Dataset("Job architecture", cols, [])

    fam_of = {a.final_profile_id: (a.final_family_id, a.final_category_id) for a in c.assignments}
    inputs: dict[int, int] = {}
    members = {g.group_id: g.member_ids for g in state.dedupe_groups}
    for a in c.assignments:
        inputs[a.final_profile_id] = inputs.get(a.final_profile_id, 0) + len(
            members.get(a.item_id, [a.item_id])
        )
    hc = _profile_headcount(state)
    je = {r.profile_key: r for r in state.je_results}
    mt = {m.profile_key: m for m in state.matching.matches}
    skills_n: dict[str, int] = {}
    for r in state.skills.profile_requirements:
        skills_n[r.profile_key] = skills_n.get(r.profile_key, 0) + 1
    tasks_n: dict[str, int] = {}
    for t in state.tasks.inferred:
        tasks_n[t.source_profile_key] = tasks_n.get(t.source_profile_key, 0) + 1

    rows = []
    for d in state.job_profiles:
        fid, cid = fam_of.get(d.profile_cluster_id, (-1, -1))
        e, m = je.get(d.profile_key), mt.get(d.profile_key)
        rows.append([
            c.family_names.get(fid, ""), c.category_names.get(cid, ""), d.title, d.profile_key,
            hc.get(d.profile_key), inputs.get(d.profile_cluster_id, 0),
            e.aggregate_score if e else None, e.level_name if e else None,
            m.family_title if m and m.matched else None,
            m.spec_title if m and m.matched else None,
            m.spec_code if m and m.matched else None,
            m.level_title if m else None,
            round(m.confidence, 2) if m else None,
            "yes" if m and m.needs_review else "",
            "yes" if m and m.overridden_by_user else "",
            skills_n.get(d.profile_key, 0), tasks_n.get(d.profile_key, 0),
            "yes" if d.stale else "",
        ])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2])))
    return Dataset("Job architecture", cols, rows)


def input_jobs_dataset(state: ProjectState) -> Dataset:
    """Every source job and where it ended up, with the full audit trail.

    This is the sheet that answers a challenge to a specific placement, so it
    carries backbone_* alongside final_* rather than only the outcome.
    """
    cols = [
        "Source job title", "Record ID", "Headcount", "Duplicate group",
        # The organisation's own structure, beside the architecture the app derived. Having
        # both on one row is what lets someone reconcile a job profile against the org
        # chart they already know — the two hierarchies do not nest, so a profile can span
        # several business units and this sheet is where that becomes visible.
        "Business level 1", "Business level 2", "Business level 3",
        "Job family", "Job category", "Job profile",
        "Stability score", "Routed by model", "Route confidence", "Moved by model",
        "Backbone profile", "Secondary profile",
    ]
    c = state.clustering
    if c is None:
        return Dataset("Input jobs", cols, [])

    titles = {r.id: r.job_title for r in state.raw_records}
    hc = {r.id: r.headcount for r in state.raw_records}
    bf = {
        r.id: (r.business_level_1 or "", r.business_level_2 or "", r.business_level_3 or "")
        for r in state.raw_records
    }
    members = {g.group_id: g.member_ids for g in state.dedupe_groups}

    rows = []
    for a in c.assignments:
        group = members.get(a.item_id, [a.item_id])
        for rec_id in group:
            rows.append([
                titles.get(rec_id, rec_id), rec_id, hc.get(rec_id),
                a.item_id if len(group) > 1 else "",
                *bf.get(rec_id, ("", "", "")),
                c.family_names.get(a.final_family_id, ""),
                c.category_names.get(a.final_category_id, ""),
                c.profile_names.get(a.final_profile_id, ""),
                round(a.stability_score, 3) if a.stability_score is not None else None,
                "yes" if a.routed_by_llm else "",
                round(a.route_confidence, 2) if a.route_confidence is not None else None,
                "yes" if a.routed_by_llm and a.backbone_profile_id != a.final_profile_id else "",
                c.profile_names.get(a.backbone_profile_id, ""),
                c.profile_names.get(a.secondary_profile_id, "")
                if a.secondary_profile_id is not None else "",
            ])
    rows.sort(key=lambda r: str(r[0]))
    return Dataset("Input jobs", cols, rows)


def skills_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Skill family", "Skill category", "Skill cluster", "Skill", "Type", "Description",
        "From job profile", "Stability score", "Routed by model",
    ]
    c = state.skills.clustering
    if c is None:
        return Dataset("Skills", cols, [])
    by_id = {s.id: s for s in state.skills.inferred}
    rows = []
    for a in c.assignments:
        s = by_id.get(a.item_id)
        if not s:
            continue
        rows.append([
            c.family_names.get(a.final_family_id, ""),
            c.category_names.get(a.final_category_id, ""),
            c.profile_names.get(a.final_profile_id, ""),
            s.name, s.kind, s.description, s.source_profile_key,
            round(a.stability_score, 3) if a.stability_score is not None else None,
            "yes" if a.routed_by_llm else "",
        ])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), str(r[3])))
    return Dataset("Skills", cols, rows)


def skill_requirements_dataset(state: ProjectState) -> Dataset:
    """The job x skill-cluster matrix with required proficiency."""
    cols = ["Job profile", "Skill cluster", "Required level", "Rationale"]
    rows = [
        [r.profile_key, r.cluster_name, r.assigned_level or "", getattr(r, "rationale", "") or ""]
        for r in state.skills.profile_requirements
    ]
    rows.sort(key=lambda r: (str(r[0]), str(r[1])))
    return Dataset("Skill requirements", cols, rows)


def proficiency_dataset(state: ProjectState) -> Dataset:
    cols = ["Skill cluster", "Level", "Definition"]
    rows = []
    for p in state.skills.cluster_proficiencies:
        for level, text in p.definitions.items():
            rows.append([p.cluster_name, level, text])
    return Dataset("Proficiency levels", cols, rows)


def tasks_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Task domain", "Task category", "Task cluster", "Task", "Description",
        "From job profile", "Time proportion %", "FTE equivalent",
        "Stability score", "Routed by model",
    ]
    c = state.tasks.clustering
    if c is None:
        return Dataset("Tasks", cols, [])
    by_id = {t.id: t for t in state.tasks.inferred}
    hc = _profile_headcount(state)
    rows = []
    for a in c.assignments:
        t = by_id.get(a.item_id)
        if not t:
            continue
        people = hc.get(t.source_profile_key)
        rows.append([
            c.family_names.get(a.final_family_id, ""),
            c.category_names.get(a.final_category_id, ""),
            c.profile_names.get(a.final_profile_id, ""),
            t.name, t.description, t.source_profile_key, t.proportion,
            round(t.proportion / 100.0 * people, 3) if people else None,
            round(a.stability_score, 3) if a.stability_score is not None else None,
            "yes" if a.routed_by_llm else "",
        ])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), -float(r[6] or 0)))
    return Dataset("Tasks", cols, rows)


def matches_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Job profile", "Matched", "Taxonomy family", "Sub-family", "Specialization",
        "Code", "Career level", "Career stream", "Cosine", "Confidence",
        "Runner-up", "Needs review", "Review reasons", "Set by user", "Rationale",
    ]
    rows = []
    for m in state.matching.matches:
        rows.append([
            m.profile_title, "yes" if m.matched else "no",
            m.family_title, m.sub_family_title, m.spec_title, m.spec_code,
            m.level_title, m.level_stream,
            round(m.cosine, 3) if m.cosine is not None else None,
            round(m.confidence, 2),
            m.runner_up_title, "yes" if m.needs_review else "",
            ", ".join(m.review_reasons), "yes" if m.overridden_by_user else "",
            m.rationale,
        ])
    rows.sort(key=lambda r: str(r[0]))
    return Dataset("Taxonomy matches", cols, rows)


def je_dataset(state: ProjectState) -> Dataset:
    """Evaluation scores with the per-domain rollups and the persona spread.

    Only the raw persona scores are persisted; the weighted totals are derived,
    so this recomputes them with the same functions the API and UI use rather
    than reimplementing the weighting and risking a different number in the
    export than on screen.

    The Generous-minus-Harsh spread is the honest measure of how contestable a
    score is, so it is a column rather than something only visible in the drawer.
    """
    from app.services.evaluation import job_evaluation as je

    # A project only stores je_framework once the user edits it; otherwise the
    # scoring ran against the shipped default. Same fallback the API uses — take
    # the wrong one and the exported domain columns are silently empty.
    framework = state.je_framework if state.je_framework.domains else je.load_default_framework()
    domains = [d.name for d in framework.domains]
    cols = [
        "Job profile", "Aggregate score", "Level",
        *je.PERSONAS, "Spread", "Stale", *domains,
    ]
    rows = []
    for r in state.je_results:
        try:
            scores = {p: je.weighted_score(r.personas[p], framework) for p in je.PERSONAS}
            subtotals = {
                d: sum(
                    (je._clip(r.personas["Balanced"][d][s.name]) - je.SCORE_MIN)
                    / (je.SCORE_MAX - je.SCORE_MIN) * 100.0 * (s.weight / 100.0)
                    for s in next(x for x in framework.domains if x.name == d).subdomains
                )
                for d in domains
            }
        except (KeyError, StopIteration):
            # The framework was edited after this result was computed; the row is
            # still worth exporting with its stored headline numbers.
            scores, subtotals = {}, {}
        g, h = scores.get("Generous"), scores.get("Harsh")
        rows.append([
            r.profile_key, r.aggregate_score, r.level_name,
            *[scores.get(p) for p in je.PERSONAS],
            round(g - h, 2) if g is not None and h is not None else None,
            "yes" if r.stale else "",
            *[round(subtotals[d], 2) if d in subtotals else None for d in domains],
        ])
    rows.sort(key=lambda r: -(r[1] or 0))
    return Dataset("Job evaluation", cols, rows)


# ---------------------------------------------------------------------------
# Work Architecture Studio
# ---------------------------------------------------------------------------
# Every one of these carries the provenance of its numbers in a column rather than in a
# footnote: the two opportunity scores are model estimates, and a spreadsheet that
# strips that context is the one that ends up pasted into a board pack.
def actions_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Task domain", "Task category", "Task cluster", "Action", "Definition",
        "% of task", "Automation % (estimate)", "Augmentation % (estimate)",
        "Cluster automation %", "Cluster augmentation %", "Scores clamped",
    ]
    c = state.tasks.clustering
    w = state.workforce
    if c is None or not w.actions:
        return Dataset("AI opportunity", cols, [])
    parents = {a.final_profile_id: (a.final_category_id, a.final_family_id) for a in c.assignments}
    rolled = {o.task_cluster_id: o for o in w.opportunity}
    rows = []
    for a in w.actions:
        cat, fam = parents.get(a.task_cluster_id, (-1, -1))
        o = rolled.get(a.task_cluster_id)
        rows.append([
            c.family_names.get(fam, ""),
            c.category_names.get(cat, ""),
            c.profile_names.get(a.task_cluster_id, ""),
            a.name, a.definition, a.pct_of_task, a.automation_pct, a.augmentation_pct,
            o.automation_pct if o else None,
            o.augmentation_pct if o else None,
            "yes" if o and o.clamped else "",
        ])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), -float(r[5] or 0)))
    return Dataset("AI opportunity", cols, rows)


def agents_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Agent", "Task cluster", "Purpose", "Automation % (estimate)",
        "Time released", "Unit", "Capabilities", "Needs a person in the loop",
        "Specification file",
    ]
    rows = [
        [
            a.name, a.cluster_name, a.purpose, a.automation_pct, a.time_released,
            a.time_released_unit, a.n_capabilities,
            "yes" if a.human_in_the_loop else "no", a.blob_path,
        ]
        for a in state.workforce.agents
    ]
    rows.sort(key=lambda r: -float(r[4] or 0))
    return Dataset("Agents", cols, rows)


def skills_guidance_dataset(state: ProjectState) -> Dataset:
    cols = ["Role", "Task cluster", "Skill file", "Description", "Hook", "Rank score", "Path"]
    rows = [
        [s.role_title, s.cluster_name, f"{s.name}.md", s.description, s.hook,
         s.rank_score, s.blob_path]
        for s in state.workforce.skills_guidance
    ]
    rows.sort(key=lambda r: (str(r[0]), -float(r[5] or 0)))
    return Dataset("Augmentation skills", cols, rows)


def processes_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Process", "Ordering confidence", "Step", "Name", "Description", "Actor",
        "System", "Already automated", "Handoff", "Sign-off", "Task cluster",
        "Match cosine", "Confirmed by model",
    ]
    rows = []
    for p in state.workforce.processes:
        for s in p.steps:
            rows.append([
                p.process_name, p.ordering_confidence, s.sequence, s.name, s.description,
                s.actor, s.system,
                "yes" if s.automated else "", "yes" if s.handoff else "",
                "yes" if s.sign_off else "",
                # The honest value for an unmatched step, not a blank that reads as
                # "not looked at": this is work no job description mentioned.
                s.task_cluster_name or "(no task cluster — not in any job description)",
                s.match_cosine if p.mapped_at else None,
                "yes" if s.routed_by_llm else "",
            ])
    return Dataset("Process steps", cols, rows)


def process_opportunity_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Process", "Steps as-is", "Steps to-be", "Manual touchpoints as-is",
        "Manual touchpoints to-be", "Actors as-is", "Actors to-be",
        "Sign-offs as-is", "Sign-offs to-be", "Handoffs as-is",
        "Handler effort reduction % (estimate)", "Elapsed time reduction % (estimate)",
        "What changes", "Risks", "Prerequisites",
    ]
    names = {p.id: p.process_name for p in state.workforce.processes}
    rows = [
        [
            names.get(a.process_id, a.process_id),
            a.as_is_steps, a.to_be_steps,
            a.as_is_manual_touchpoints, a.to_be_manual_touchpoints,
            a.as_is_actors, a.to_be_actors,
            a.as_is_sign_offs, a.to_be_sign_offs, a.as_is_handoffs,
            a.effort_reduction_pct, a.elapsed_reduction_pct,
            " | ".join(a.what_changes), " | ".join(a.risks), " | ".join(a.prerequisites),
        ]
        for a in state.workforce.process_assessments
    ]
    return Dataset("Process opportunity", cols, rows)


def future_roles_dataset(state: ProjectState) -> Dataset:
    cols = [
        "Role", "Automation % (estimate)", "% of week changing shape", "Future purpose",
        "Today", "First to change", "What it becomes", "Future responsibilities",
        "Changes shape first", "Deepens", "Skills to build", "Keep sharp by hand",
    ]
    rows = [
        [
            f.title, f.automation_pct, f.time_released_pct, f.future_purpose,
            f.evolution_today, f.evolution_after_automation, f.evolution_future,
            " | ".join(f.future_responsibilities), " | ".join(f.absorbed_tasks),
            " | ".join(f.deepened_tasks), " | ".join(f.skills_to_build),
            " | ".join(f.deliberate_practice),
        ]
        for f in state.workforce.future_roles
    ]
    rows.sort(key=lambda r: -float(r[2] or 0))
    return Dataset("Future roles", cols, rows)


BUILDERS = {
    "architecture": architecture_dataset,
    "input-jobs": input_jobs_dataset,
    "evaluation": je_dataset,
    "skills": skills_dataset,
    "skill-requirements": skill_requirements_dataset,
    "proficiency": proficiency_dataset,
    "tasks": tasks_dataset,
    "matches": matches_dataset,
    "ai-opportunity": actions_dataset,
    "agents": agents_dataset,
    "augmentation-skills": skills_guidance_dataset,
    "processes": processes_dataset,
    "process-opportunity": process_opportunity_dataset,
    "future-roles": future_roles_dataset,
}


def build_all(state: ProjectState) -> list[Dataset]:
    """Every dataset that has rows. An empty sheet in a client deliverable reads
    as a mistake, so stages that haven't run are omitted rather than blank."""
    out = []
    for build in BUILDERS.values():
        ds = build(state)
        if ds.rows:
            out.append(ds)
    return out
